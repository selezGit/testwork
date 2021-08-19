import os
from datetime import datetime

from openpyxl.styles import NamedStyle, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.fonts import Font

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BASE_URL = (
    'https://www.moex.com/export/derivatives/currency-rate.aspx?language=ru'
)
FILE_NAME = f'report-{datetime.today().strftime("%Y-%m-%d %H:%M:%S")}.xlsx'
PARSE_LIST = ['USD/RUB', 'EUR/RUB']

# openpyxl
HEADER_FILL = PatternFill(
    start_color='4169E1',
    end_color='4169E1',
    fill_type='solid',
)
BASE_FILL = PatternFill(
    start_color='E0FFFF',
    end_color='E0FFFF',
    fill_type='solid',
)
RED_FILL = PatternFill(
    start_color='F08080',
    end_color='F08080',
    fill_type='solid',
)
GREEN_FILL = PatternFill(
    start_color='C2D69B',
    end_color='C2D69B',
    fill_type='solid',
)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)
HEADER_STYLE = NamedStyle(
    name='header',
    border=THIN_BORDER,
    fill=HEADER_FILL,
    font=Font(color='00FFFFFF'),
)
BASE_STYLE = NamedStyle(
    name='base',
    border=THIN_BORDER,
    fill=BASE_FILL,
)
RED_STYLE = NamedStyle(
    name='red',
    border=THIN_BORDER,
    fill=RED_FILL,
)
GREEN_STYLE = NamedStyle(
    name='green',
    border=THIN_BORDER,
    fill=GREEN_FILL,
)
NUMBER_FORMAT = u'#,####0.0000;_(* (-#,###0.0000);_(* -_0_0_)'
RIGHT_INDENT = 1.15

# email
SMTP_SERVER = os.getenv('SMTP_SERVER')
SMTP_PORT = os.getenv('SMTP_PORT', 465)
EMAIL_LOGIN = os.getenv('EMAIL_LOGIN')
EMAIL_PASSWORD = os.getenv('EMAIL_PASSWORD')
