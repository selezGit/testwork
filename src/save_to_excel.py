from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from src.core import config


def save_to_excel(
    cleared_dict: dict,
    offset: int,
    exchange: str,
) -> None:
    try:
        wb = load_workbook(config.FILE_NAME)
    except FileNotFoundError:
        wb = Workbook()

    ws = wb.active

    if 'EUR' in exchange:
        currency = f'€{config.NUMBER_FORMAT}'
    else:
        currency = f'${config.NUMBER_FORMAT}'

    ws.cell(
        column=offset + 1,
        row=1,
        value='Дата',
    ).style = config.HEADER_STYLE
    ws.cell(
        column=offset + 2,
        row=1,
        value='Курс',
    ).style = config.HEADER_STYLE

    ws.cell(
        column=offset + 3,
        row=1,
        value='Изменение',
    ).style = config.HEADER_STYLE

    for row, date in enumerate(cleared_dict.keys()):
        try:
            exchange_rate = float(cleared_dict[date][0]['value'])
            difference = exchange_rate - float(cleared_dict[date][1]['value'])
            color = config.RED if difference <= 0 else config.GREEN
        except ValueError:
            exchange_rate = float(cleared_dict[date][1]['value'])
            difference = 0
            color = config.BASE
        except IndexError:
            exchange_rate = float(cleared_dict[date][0]['value'])
            difference = 0
            color = config.BASE

        ws.cell(
            column=offset + 1,
            row=row + 2,
            value=date,
        ).style = config.BASE_STYLE

        cell_rate = ws.cell(
            column=offset + 2,
            row=row + 2,
            value=exchange_rate,
        )
        cell_rate.style = config.BASE_STYLE
        cell_rate.number_format = currency

        cell_diff = ws.cell(
            column=offset + 3,
            row=row + 2,
            value=difference,
        )
        cell_diff.number_format = currency
        cell_diff.border = config.THIN_BORDER
        cell_diff.fill = color

    wb.save(config.FILE_NAME)


def update_data() -> int:
    wb = load_workbook(config.FILE_NAME)
    ws = wb.active
    for index, row in enumerate(ws.iter_rows()):
        if index == 0:
            continue
        usd, eur = 0, 0
        for cell in row:
            if cell.column == 2:
                usd = cell.value
            elif cell.column == 5:
                eur = cell.value
        try:
            mid_market_rate = eur / usd
        except ZeroDivisionError:
            mid_market_rate = 0

        res_col = ws.cell(
            column=7,
            row=index + 1,
            value=mid_market_rate,
        )
        res_col.number_format = config.NUMBER_FORMAT
        res_col.style = config.BASE_STYLE

    ws.cell(column=7, row=1, value='Средний курс').style = config.HEADER_STYLE

    # авто выравнивание
    for i in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].bestFit = True
        ws.column_dimensions[get_column_letter(i)].auto_size = True

    wb.save(config.FILE_NAME)

    return index + 1
