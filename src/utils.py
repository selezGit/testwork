from datetime import datetime
from functools import wraps
from typing import Any, Callable, List, Tuple

import pymorphy2
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook, load_workbook

from src.core import config


def get_dates() -> Tuple[str, str]:
    current_date = datetime.today()
    past_date = current_date - relativedelta(months=1)
    return current_date.date(), past_date.date()


def format_data(raw_data: List[dict]) -> dict:
    cleared_dict = {}
    for raw_dict in raw_data:
        date_string, time_string = raw_dict['@moment'].split(' ')
        cleared_dict.setdefault(date_string, []).append(
            {'time': time_string, 'value': raw_dict['@value']}
        )

    return cleared_dict


def get_morphy(colls_count: int) -> str:
    morph = pymorphy2.MorphAnalyzer()
    string = morph.parse('строка')[0]
    return string.make_agree_with_number(colls_count).word


def connect_to_wb(func: Callable):
    @wraps(func)
    def inner(*args: Any, **kwargs: Any):
        try:
            wb = load_workbook(config.FILE_NAME)
        except FileNotFoundError:
            wb = Workbook()
        ws = wb.active
        result = func(ws, *args, **kwargs)
        wb.save(config.FILE_NAME)

        return result
    return inner
